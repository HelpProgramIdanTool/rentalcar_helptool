import re
import unicodedata
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from suppliers.models import Supplier, SupplierLocation


AIRPORT_CODES = {
    "Bydgoszcz": "BZG",
    "Gdańsk": "GDN",
    "Katowice Pyrzowice": "KTW",
    "Kraków Balice": "KRK",
    "Lublin": "LUZ",
    "Łódź": "LCJ",
    "Modlin": "WMI",
    "Olsztyn Szymany": "SZY",
    "Pardubice Airport": "PED",
    "Poznań": "POZ",
    "Prague Airport": "PRG",
    "Radom": "RDO",
    "Rzeszów": "RZE",
    "Szczecin Goleniów": "SZZ",
    "Warszawa-Chopin": "WAW",
    "Wrocław": "WRO",
}

CZECH_LOCATIONS = {"Pardubice Airport", "Prague Airport"}

POLISH_LETTERS = str.maketrans({"ł": "l", "Ł": "L"})


def make_code(value):
    value = value.translate(POLISH_LETTERS)
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]+", "-", value.upper()).strip("-")


def parse_carfree_locations(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["CarFree Departments"]
    except KeyError as error:
        raise CommandError("Sheet 'CarFree Departments' was not found.") from error

    locations = []
    used_codes = set()

    for city, source_type, address, *_ in worksheet.iter_rows(values_only=True):
        if not city or not source_type:
            continue
        if str(city).strip().lower() == "city":
            continue

        city = str(city).strip()
        source_type = str(source_type).strip()
        address = str(address or "").strip()
        source_type_lower = source_type.lower()
        is_airport = "airport" in source_type_lower or "airport" in city.lower()
        is_meet_and_greet = "meet & greet" in source_type_lower

        if is_airport:
            location_type = SupplierLocation.LocationType.AIRPORT
            airport_code = AIRPORT_CODES.get(city, "")
            base_code = airport_code or f"{make_code(city)}-AIRPORT"
        elif "bus station" in source_type_lower:
            location_type = SupplierLocation.LocationType.BRANCH
            airport_code = ""
            base_code = f"{make_code(city)}-BUS"
        elif "train station" in source_type_lower:
            location_type = SupplierLocation.LocationType.BRANCH
            airport_code = ""
            base_code = f"{make_code(city)}-TRAIN"
        else:
            location_type = SupplierLocation.LocationType.BRANCH
            airport_code = ""
            base_code = f"{make_code(city)}-CITY"

        location_code = base_code
        number = 2
        while location_code in used_codes:
            location_code = f"{base_code}-{number}"
            number += 1
        used_codes.add(location_code)

        locations.append(
            {
                "location_code": location_code,
                "location_name": city,
                "city": city,
                "country": (
                    "Czech Republic"
                    if city in CZECH_LOCATIONS or "czech republic" in address.lower()
                    else "Poland"
                ),
                "address": address,
                "location_type": location_type,
                "airport_code": airport_code,
                "supports_pickup": True,
                "supports_return": True,
                "has_rental_desk": not is_meet_and_greet,
                "supports_terminal_delivery": is_meet_and_greet,
                "supports_address_delivery": False,
                "is_active": True,
            }
        )

    workbook.close()
    return locations


class Command(BaseCommand):
    help = "Preview or import CarFree locations from its price-list workbook."

    def add_arguments(self, parser):
        parser.add_argument("file", type=Path)
        parser.add_argument("--supplier-code", default="03")
        parser.add_argument("--preview", action="store_true")

    def handle(self, *args, **options):
        file_path = options["file"]
        if not file_path.is_file():
            raise CommandError(f"File not found: {file_path}")

        locations = parse_carfree_locations(file_path)
        if not locations:
            raise CommandError("No locations were found.")

        for location in locations:
            service = "terminal delivery" if location["supports_terminal_delivery"] else "rental desk"
            self.stdout.write(
                f'{location["location_code"]}: {location["location_name"]} | '
                f'{location["location_type"]} | {service} | {location["country"]}'
            )

        if options["preview"]:
            self.stdout.write(self.style.WARNING(f"Preview only: {len(locations)} locations, nothing saved."))
            return

        try:
            supplier = Supplier.objects.get(supplier_code=options["supplier_code"])
        except Supplier.DoesNotExist as error:
            raise CommandError(
                f'Supplier with code {options["supplier_code"]!r} was not found.'
            ) from error

        created_count = 0
        updated_count = 0
        with transaction.atomic():
            for location in locations:
                location_code = location.pop("location_code")
                _, created = SupplierLocation.objects.update_or_create(
                    supplier=supplier,
                    location_code=location_code,
                    defaults=location,
                )
                if created:
                    created_count += 1
                else:
                    updated_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete: {created_count} created, {updated_count} updated."
            )
        )
