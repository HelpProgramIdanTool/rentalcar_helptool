from datetime import date
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from suppliers.models import Supplier, SupplierExtra, SupplierExtraRate


PRICE_LIST_DIR = Path("supplyer price lists")


def rate(
    amount,
    calculation_type,
    *,
    rate_code="DEFAULT",
    maximum=None,
    minimum=None,
    days_from=None,
    days_to=None,
    formula=None,
):
    return {
        "rate_code": rate_code,
        "calculation_type": calculation_type,
        "amount_gross": Decimal(str(amount)),
        "currency": "PLN",
        "minimum_amount_gross": Decimal(str(minimum)) if minimum is not None else None,
        "maximum_amount_gross": Decimal(str(maximum)) if maximum is not None else None,
        "days_from": days_from,
        "days_to": days_to,
        "formula_config": formula or {},
    }


def extra(code, name, category, source_label, rates, description=""):
    return {
        "extra_code": code,
        "name": name,
        "category": category,
        "source_label": source_label,
        "description": description,
        "rates": rates,
    }


def read_two_column_sheet(path, sheet_name, label_column, price_column):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    values = {}
    for row in sheet.iter_rows(values_only=True):
        label = row[label_column] if len(row) > label_column else None
        price = row[price_column] if len(row) > price_column else None
        if label and price is not None:
            values[str(label).strip()] = str(price).strip()
    return values


def car_free_records():
    fixed = SupplierExtraRate.CalculationType.FIXED
    rental = SupplierExtraRate.CalculationType.PER_RENTAL
    unit = SupplierExtraRate.CalculationType.PER_UNIT
    day = SupplierExtraRate.CalculationType.PER_DAY
    formula = SupplierExtraRate.CalculationType.FORMULA
    return [
        extra("KEY_DAMAGE", "Damaged or misplaced key", "PENALTY", "Damage or misplacement of a car key or remote control", [rate(1000, formula, formula={"plus": "dealer repair cost"})]),
        extra("KEY_LOSS", "Lost car key", "PENALTY", "Loss of a car key or remote control", [rate(5000, fixed)]),
        extra("FINE_ADMIN_PL", "Polish fine administration", "PENALTY", "Administrative fee in connection with the handling of fine payment obligations to Polish institutions", [rate(100, formula, formula={"plus": "actual fine"})]),
        extra("THIRD_PARTY_INQUIRY", "Third-party inquiry administration", "PENALTY", "Administrative fee for handling third-party inquiries related to the use of the vehicle during the rental period", [rate(100, formula, formula={"plus": "actual fine"})]),
        extra("FINE_ADMIN_FOREIGN", "Foreign fine administration", "PENALTY", "Administrative fee in connection with the handling of fine payment obligations to foreign institutions", [rate(250, formula, formula={"plus": "actual fine"})]),
        extra("YOUNG_DRIVER_21_24", "Young driver age 21-24", "DRIVER", "Young Driver's fee (21-24)", [rate(0, rental)]),
        extra("ONE_WAY_PL", "One-way rental in Poland", "DELIVERY", "One way in Poland (up to 3 days, 4 days and above is free)", [rate(199, rental, rate_code="DAYS_1_3", days_from=1, days_to=3), rate(0, rental, rate_code="DAYS_4_PLUS", days_from=4)]),
        extra("BABY_SEAT_BOOSTER", "Baby seat or booster", "CHILD_EQUIPMENT", "Baby Seat/Booster (up to 10 days, 11 days and above is free)", [rate(25, day, maximum=250)]),
        extra("FOREIGN_CITY_DELIVERY", "Delivery or return in selected foreign city", "DELIVERY", "Delivery / Return in Bratislava, Berlin, Vienna, Budapest", [rate(1000, unit)]),
        extra("LOST_PARKING_TICKET", "Lost parking ticket", "PENALTY", "No parking ticket", [rate(100, formula, formula={"plus": "actual parking fee"})]),
        extra("CITY_ADDRESS_DELIVERY", "Address delivery or return in branch city", "DELIVERY", "Delivery/Return to an address in a city where CarFree is located", [rate(59, unit)], "Up to 30 km from the branch."),
        extra("OUTSIDE_CITY_DELIVERY", "Address delivery or return outside branch city", "DELIVERY", "Delivery/Return to an address in a city without a CarFree branch", [rate(0, formula, formula={"per_km_gross": "3.00"})]),
        extra("OUT_OF_HOURS", "Out-of-hours pickup or return", "AFTER_HOURS", "Out of hours pick up / return", [rate(60, unit)]),
        extra(
            "CROSS_BORDER",
            "Travel abroad",
            "CROSS_BORDER",
            "Travel outside the country (Germany, Czech Republic, Slovakia, Hungary, Lithuania, Latvia, Estonia, Austria, Denmark, France, Italy, Croatia, Romania, Bulgaria, Switzerland, Slovenia, Netherlands, Belgium, Spain, Portugal, Ireland, United Kingdom, Norway, Sweden, Greece)",
            [
                rate(
                    299,
                    formula,
                    formula={
                        "per_rental_gross": "299.00",
                        "per_rental_day_gross": "30.00",
                        "clarification": "Confirmed by Idan",
                    },
                )
            ],
            "Customer price confirmed by Idan: 299 PLN plus 30 PLN for every rental day.",
        ),
        extra("UNAUTHORISED_CROSS_BORDER", "Unauthorised travel abroad", "PENALTY", "Unauthorised Travel Abroad", [rate(5000, fixed)]),
        extra("ADDITIONAL_DRIVER", "Additional driver", "DRIVER", "Additional driver charge", [rate(15, day)]),
        extra("IDP_MISSING", "Missing international driving permit", "DRIVER", "International Drivers permit required", [rate(50, day)]),
        extra("MISSING_FUEL", "Missing fuel", "FUEL", "Fuel Policy Charge", [rate(100, formula, formula={"per_missing_liter_gross": "10.00"})]),
    ]


def kaizen_records():
    rental = SupplierExtraRate.CalculationType.PER_RENTAL
    unit = SupplierExtraRate.CalculationType.PER_UNIT
    day = SupplierExtraRate.CalculationType.PER_DAY
    driver_day = SupplierExtraRate.CalculationType.PER_DRIVER_DAY
    fixed = SupplierExtraRate.CalculationType.FIXED
    formula = SupplierExtraRate.CalculationType.FORMULA
    return [
        extra("OFFICE_AIRPORT_SERVICE", "Office or airport delivery and pickup", "DELIVERY", "Delivery and pickup at the airports or cities where are Kaizen Rent offices", [rate(0, unit)]),
        extra("CITY_ADDRESS_DELIVERY", "Delivery and pickup at customer address in office city", "DELIVERY", "Delivery and pickup in the client's location in the city where Kaizen Rent has the office", [rate(100, unit)]),
        extra("OUTSIDE_CITY_DELIVERY", "Delivery and pickup outside office city", "DELIVERY", "Delivery and pickup in the client's location outside the city where Kaizen Rent has the office", [rate(70, formula, formula={"per_km_gross": "1.00"})]),
        extra("NIGHT_SERVICE", "Service between 20:00 and 08:00", "AFTER_HOURS", "Service in 20:00-8:00", [rate(70, unit)]),
        extra("AIRPORT_24_7", "24/7 service at selected airports", "AFTER_HOURS", "Service 24/7 in Airports: Gdańsk, Katowice, Kraków, Warszawa", [rate(0, unit)]),
        extra("CROSS_BORDER", "Travel abroad", "CROSS_BORDER", "Trip abroad (per rental)", [rate(299, rental)]),
        extra("CROSS_BORDER_PREMIUM", "Travel abroad - Premium", "CROSS_BORDER", "Trip abroad Premium (per rental)", [rate(499, rental)]),
        extra("YOUNG_DRIVER", "Young driver below 24", "DRIVER", "Young driver fee below 24 years", [rate("29.99", driver_day)]),
        extra("ADDITIONAL_DRIVER", "Additional driver", "DRIVER", "Additional driver  (per rental)", [rate(0, rental)]),
        extra("ONE_WAY", "Return in a different location", "DELIVERY", "One way - pick up and collection in a different locations", [rate(0, rental)]),
        extra("SHORT_RENTAL", "Short rental fee", "RENTAL", "Short rental fee (1-2 day rental) - ADDITIONAL PRICE PER DAY", [rate(0, day, days_from=1, days_to=2)]),
        extra("NAVIGATION", "Navigation", "EQUIPMENT", "Navigation (per rental)", [rate(10, day, maximum=150)]),
        extra("DAMAGE_SHARE", "Share in damage", "INSURANCE", "Share in damage", [rate(4000, fixed)]),
        extra("FINAL_WASH", "Final car washing", "CLEANING", "Final car washing", [rate(0, rental)]),
        extra("FINAL_REFUELLING", "Final refuelling", "FUEL", "Final refuelling of the car", [rate(340, fixed)]),
        extra("CHILD_SEAT", "Child seat", "CHILD_EQUIPMENT", "Child seat (per day)", [rate(20, day, maximum=200)]),
        extra("SNOW_CHAINS", "Snow chains", "EQUIPMENT", "Chains for wheels (per day)", [rate(10, day)]),
        extra("SMOKE_CLEANING", "Cleaning after smoking", "PENALTY", "Ticket for back with the smoke smell", [rate(500, fixed)]),
        extra("UPHOLSTERY_CLEANING", "Upholstery cleaning", "PENALTY", "Ticket for wash of upholstery", [rate(500, fixed)]),
        extra("WIFI_ROUTER", "Wi-Fi router", "EQUIPMENT", "WIFI Router", [rate(20, day, maximum=200)]),
        extra("VIP_SERVICE", "VIP service", "SERVICE", "VIP service (50% commission)", [rate(200, rental)]),
    ]


def one_rent_records():
    rental = SupplierExtraRate.CalculationType.PER_RENTAL
    unit = SupplierExtraRate.CalculationType.PER_UNIT
    day = SupplierExtraRate.CalculationType.PER_DAY
    fixed = SupplierExtraRate.CalculationType.FIXED
    formula = SupplierExtraRate.CalculationType.FORMULA
    shared_equipment = "CHILD SEAT, GPS, SNOW CHAINS"
    return [
        extra("CITY_AIRPORT_DELIVERY", "Delivery or return in city or airport", "DELIVERY", "DELIVERY/RETURN in city or airport", [rate(100, unit)]),
        extra("OUTSIDE_CITY_DELIVERY", "Delivery or return outside the city", "DELIVERY", "DELIVERY/RETURN outside the city", [rate(100, formula, formula={"per_km_gross": "2.50"})]),
        extra("AIRPORT_FEE", "Airport fee", "AIRPORT", "AIRPORT FEE", [rate(50, rental)]),
        extra("ONE_WAY", "Return in another location", "DELIVERY", "RETURN IN OTHER LOCATION", [rate(0, rental)]),
        extra("MISSING_FUEL", "Missing fuel", "FUEL", "LACK OF PETROL", [rate(100, formula, formula={"plus": "actual fuel cost"})]),
        extra("CHILD_SEAT", "Child seat", "CHILD_EQUIPMENT", shared_equipment, [rate(20, day, maximum=150)]),
        extra("GPS", "GPS navigation", "EQUIPMENT", shared_equipment, [rate(20, day, maximum=150)]),
        extra("SNOW_CHAINS", "Snow chains", "EQUIPMENT", shared_equipment, [rate(20, day, maximum=150)]),
        extra("ADDITIONAL_DRIVER", "Additional driver", "DRIVER", "ADDITIONAL DRIVERS", [rate(0, rental)]),
        extra("CROSS_BORDER", "Travel abroad", "CROSS_BORDER", "ACCESS TO TRAVEL ABROAD", [rate(200, rental)]),
        extra("EXTRA_CLEANING", "Extra cleaning", "CLEANING", "EXTRA CLEANING", [rate(300, fixed)]),
        extra("DAMAGE_EXCESS_FEE", "Damage excess fee", "INSURANCE", "FEE DAMAGE ACCESS", [rate(0, fixed)]),
    ]


def upsert_supplier_extras(supplier, records, source_values, valid_from, source_name):
    extras_created = rates_created = 0
    for record in records:
        label = record["source_label"]
        if label not in source_values:
            raise CommandError(f"Missing source row in {source_name}: {label}")
        raw_price = source_values[label]
        description = record["description"]
        source_note = f"Source: {source_name}. Original price: {raw_price}"
        if description:
            description = f"{description}\n{source_note}"
        else:
            description = source_note
        extra_obj, created = SupplierExtra.objects.update_or_create(
            supplier=supplier,
            extra_code=record["extra_code"],
            defaults={
                "name": record["name"],
                "category": record["category"],
                "description": description,
                "is_active": True,
            },
        )
        extras_created += int(created)
        for rate_data in record["rates"]:
            rate_code = rate_data.pop("rate_code")
            _, rate_created = SupplierExtraRate.objects.update_or_create(
                extra=extra_obj,
                rate_code=rate_code,
                defaults={**rate_data, "valid_from": valid_from, "is_active": True, "note": raw_price},
            )
            rate_data["rate_code"] = rate_code
            rates_created += int(rate_created)
    return extras_created, rates_created


class Command(BaseCommand):
    help = "Import supplier extras and customer-facing gross prices from Excel price lists."

    @transaction.atomic
    def handle(self, *args, **options):
        sources = [
            (
                "Car Free",
                PRICE_LIST_DIR / "CarFree 13.08.2026.xlsx",
                "Fees",
                0,
                1,
                car_free_records(),
                date(2026, 8, 13),
            ),
            (
                "Kaizen Rent",
                PRICE_LIST_DIR / "Kaizen Rent.xlsx",
                "Additional fees",
                0,
                1,
                kaizen_records(),
                date(2026, 6, 24),
            ),
            (
                "One Rent",
                PRICE_LIST_DIR / "One Rent 2026.xlsx",
                "High season",
                1,
                4,
                one_rent_records(),
                date(2026, 1, 1),
            ),
        ]
        total_extras = total_rates = 0
        for supplier_name, path, sheet, label_col, price_col, records, valid_from in sources:
            if not path.exists():
                raise CommandError(f"Price list not found: {path}")
            try:
                supplier = Supplier.objects.get(supplier_name=supplier_name)
            except Supplier.DoesNotExist as error:
                raise CommandError(f"Supplier not found: {supplier_name}") from error
            source_values = read_two_column_sheet(path, sheet, label_col, price_col)
            created_extras, created_rates = upsert_supplier_extras(
                supplier,
                records,
                source_values,
                valid_from,
                path.name,
            )
            total_extras += created_extras
            total_rates += created_rates
            self.stdout.write(
                f"{supplier_name}: {len(records)} extras, "
                f"{sum(len(item['rates']) for item in records)} rates"
            )
        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete. Created {total_extras} extras and {total_rates} rates."
            )
        )
