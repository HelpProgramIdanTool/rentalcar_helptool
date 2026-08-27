import re
from collections import OrderedDict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from suppliers.management.commands.import_carfree_locations import make_code
from suppliers.models import Supplier, VehicleGroup, VehicleModel


BODY_TYPE_BY_ACRISS = {
    "D": VehicleGroup.BodyType.HATCHBACK,
    "F": VehicleGroup.BodyType.SUV,
    "G": VehicleGroup.BodyType.SUV,
    "K": VehicleGroup.BodyType.VAN,
    "L": VehicleGroup.BodyType.SEDAN,
    "P": VehicleGroup.BodyType.PICKUP,
    "V": VehicleGroup.BodyType.MINIVAN,
    "W": VehicleGroup.BodyType.ESTATE,
}


def transmission_from_acriss(code):
    if len(code) < 3:
        return VehicleGroup.Transmission.UNKNOWN
    if code[2] in "ABD":
        return VehicleGroup.Transmission.AUTOMATIC
    if code[2] in "CMN":
        return VehicleGroup.Transmission.MANUAL
    return VehicleGroup.Transmission.UNKNOWN


def body_type_from_acriss(code):
    if len(code) < 2:
        return ""
    return BODY_TYPE_BY_ACRISS.get(code[1], "")


def split_brand_and_model(value):
    value = re.sub(r"\s+", " ", value.strip())
    value = re.sub(r"\s+-\s+\d+$", "", value)
    value = value.replace("Hyunda i30", "Hyundai i30")
    if value.upper().startswith("BMW"):
        model = value[3:].strip() or value
        return "BMW", model
    if value.upper().startswith("VW "):
        return "Volkswagen", value[3:].strip()
    parts = value.split(" ", 1)
    if len(parts) == 1:
        return "", value
    known_brands = {
        "Audi", "Citroen", "Cupra", "Fiat", "Ford", "Hyundai", "Isuzu",
        "Kia", "Lexus", "Mazda", "Mercedes", "Nissan", "Opel", "Peugeot",
        "Renault", "Seat", "Skoda", "Toyota", "Volkswagen", "Volvo",
    }
    brand = parts[0].title()
    return (brand, parts[1]) if brand in known_brands else ("", value)


def split_models(value, with_counts=False):
    models = []
    for item in str(value or "").split(","):
        item = item.strip()
        if not item:
            continue
        if with_counts:
            item = re.sub(r"\s+-\s+\d+$", "", item)
        models.append(split_brand_and_model(item))
    return models


def group_name_for_kaizen(source_name, code):
    if source_name == "C Aut" and code == "CDAR":
        return "C Automatic Hatchback"
    if source_name == "C Aut" and code == "CLAR":
        return "C Automatic Sedan"
    if source_name == "SUV Premium" and code == "PFAH":
        return "SUV Premium Hybrid"
    return source_name


def add_record(records, code, name, models, display_order, transmission=None, body_type=None):
    record = records.setdefault(
        code,
        {
            "group_code": code,
            "group_name": name,
            "transmission": transmission or transmission_from_acriss(code),
            "body_type": body_type if body_type is not None else body_type_from_acriss(code),
            "display_order": display_order,
            "models": [],
        },
    )
    for model in models:
        brand, model_name = model
        matching_index = next(
            (
                index
                for index, existing in enumerate(record["models"])
                if existing[1].casefold() == model_name.casefold()
            ),
            None,
        )
        if matching_index is None:
            record["models"].append(model)
        elif not record["models"][matching_index][0] and brand:
            record["models"][matching_index] = model


def read_kaizen(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook["Idan"]
    records = OrderedDict()
    order = 1
    for row in worksheet.iter_rows(values_only=True):
        if len(row) < 4 or not all(row[index] for index in (0, 1, 2)):
            continue
        source_name = str(row[0]).strip()
        source_codes = str(row[1]).strip()
        if source_name in {"Class", "IDAN!!!"} or source_codes in {"Acriss", "brutto"}:
            continue
        codes = [code.strip() for code in source_codes.split(";") if code.strip()]
        fleet_models = split_models(row[3], with_counts=True)
        example_model = split_brand_and_model(str(row[2]))
        fleet_keys = {tuple(part.casefold() for part in model) for model in fleet_models}
        if tuple(part.casefold() for part in example_model) not in fleet_keys:
            fleet_models.append(example_model)

        for code in codes:
            code_models = fleet_models
            if source_name == "C Aut" and len(codes) > 1:
                if code == "CLAR":
                    code_models = [model for model in fleet_models if model[1].lower() == "corolla"]
                elif code == "CDAR":
                    code_models = [model for model in fleet_models if model[1].lower() != "corolla"]
            add_record(
                records,
                code,
                group_name_for_kaizen(source_name, code),
                code_models,
                order,
            )
            order += 1
    workbook.close()
    return list(records.values())


def read_one_rent(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook["High season"]
    records = OrderedDict()
    order = 1
    for row in worksheet.iter_rows(values_only=True):
        if len(row) < 4 or not all(row[index] for index in (1, 2, 3)):
            continue
        name = str(row[1]).strip()
        code = str(row[2]).strip()
        if code == "ACRISS/SIPP":
            continue
        add_record(records, code, name, split_models(row[3]), order)
        order += 1
    workbook.close()
    return list(records.values())


def carfree_body_type(segment):
    value = segment.lower()
    if "suv" in value or "crossover" in value:
        return VehicleGroup.BodyType.SUV
    if "station wagon" in value:
        return VehicleGroup.BodyType.ESTATE
    if "bus" in value:
        return VehicleGroup.BodyType.VAN
    return ""


def read_carfree(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook["CarFree Fleet"]
    records = OrderedDict()
    order = 1
    for row in worksheet.iter_rows(values_only=True):
        if len(row) < 3 or not all(row[index] for index in (0, 1, 2)):
            continue
        segment = str(row[0]).strip()
        transmission_name = str(row[1]).strip()
        if segment == "Segment":
            continue
        code = f"{make_code(segment)}-{make_code(transmission_name)}"
        transmission = (
            VehicleGroup.Transmission.AUTOMATIC
            if transmission_name.lower() == "automatic"
            else VehicleGroup.Transmission.MANUAL
        )
        add_record(
            records,
            code,
            f"{segment} {transmission_name}",
            [split_brand_and_model(str(row[2]))],
            order,
            transmission=transmission,
            body_type=carfree_body_type(segment),
        )
        order += 1
    workbook.close()
    return list(records.values())


class Command(BaseCommand):
    help = "Preview or import vehicle groups and example models from supplier workbooks."

    def add_arguments(self, parser):
        parser.add_argument("directory", type=Path)
        parser.add_argument("--preview", action="store_true")

    def handle(self, *args, **options):
        directory = options["directory"]
        sources = [
            ("01", "Kaizen Rent.xlsx", read_kaizen),
            ("02", "One Rent 2026.xlsx", read_one_rent),
            ("03", "CarFree 13.08.2026.xlsx", read_carfree),
        ]
        parsed = []
        for supplier_code, filename, reader in sources:
            file_path = directory / filename
            if not file_path.is_file():
                raise CommandError(f"File not found: {file_path}")
            records = reader(file_path)
            parsed.append((supplier_code, filename, records))
            self.stdout.write(f"{filename}: {len(records)} groups")
            for record in records:
                self.stdout.write(
                    f'  {record["group_code"]}: {record["group_name"]} | '
                    f'{record["transmission"]} | models={len(record["models"])}'
                )

        total = sum(len(records) for _, _, records in parsed)
        if options["preview"]:
            self.stdout.write(
                self.style.WARNING(f"Preview only: {total} groups, nothing saved.")
            )
            return

        created_groups = 0
        updated_groups = 0
        created_models = 0
        with transaction.atomic():
            for supplier_code, _, records in parsed:
                try:
                    supplier = Supplier.objects.get(supplier_code=supplier_code)
                except Supplier.DoesNotExist as error:
                    raise CommandError(
                        f"Supplier with code {supplier_code!r} was not found."
                    ) from error
                for record in records:
                    group, created = VehicleGroup.objects.update_or_create(
                        supplier=supplier,
                        group_code=record["group_code"],
                        defaults={
                            "group_name": record["group_name"],
                            "transmission": record["transmission"],
                            "body_type": record["body_type"],
                            "display_order": record["display_order"],
                            "is_active": True,
                        },
                    )
                    if created:
                        created_groups += 1
                    else:
                        updated_groups += 1
                    for index, (brand, model) in enumerate(record["models"], 1):
                        _, model_created = VehicleModel.objects.update_or_create(
                            vehicle_group=group,
                            brand=brand,
                            model=model,
                            defaults={"display_order": index, "is_active": True},
                        )
                        if model_created:
                            created_models += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete: {created_groups} groups created, "
                f"{updated_groups} groups updated, {created_models} models created."
            )
        )
