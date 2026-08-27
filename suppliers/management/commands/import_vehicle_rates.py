from datetime import date
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from suppliers.models import (
    PriceDayRange,
    PriceList,
    PriceSeason,
    Supplier,
    VehicleGroup,
    VehicleRate,
)


PRICE_LIST_DIR = Path("supplyer price lists")


def get_supplier(name):
    try:
        return Supplier.objects.get(supplier_name=name)
    except Supplier.DoesNotExist as error:
        raise CommandError(f"Supplier not found: {name}") from error


def get_group(supplier, code):
    try:
        return supplier.vehicle_groups.get(group_code=code)
    except VehicleGroup.DoesNotExist as error:
        raise CommandError(
            f"Vehicle group {code} not found for {supplier.supplier_name}"
        ) from error


def upsert_price_list(supplier, name, version, source_file, effective_from, effective_to=None, note=""):
    price_list, _ = PriceList.objects.update_or_create(
        supplier=supplier,
        version=version,
        defaults={
            "name": name,
            "effective_from": effective_from,
            "effective_to": effective_to,
            "currency": "PLN",
            "status": PriceList.Status.ACTIVE,
            "source_type": PriceList.SourceType.EXCEL,
            "source_file": source_file,
            "note": note,
        },
    )
    return price_list


def upsert_season(price_list, code, name, start, end, note=""):
    season, _ = PriceSeason.objects.update_or_create(
        price_list=price_list,
        season_code=code,
        defaults={
            "season_name": name,
            "rental_date_from": start,
            "rental_date_to": end,
            "priority": 0,
            "is_active": True,
            "note": note,
        },
    )
    return season


def upsert_ranges(price_list, definitions):
    ranges = []
    for order, (code, label, days_from, days_to) in enumerate(definitions, 1):
        day_range, _ = PriceDayRange.objects.update_or_create(
            price_list=price_list,
            range_code=code,
            defaults={
                "label": label,
                "days_from": days_from,
                "days_to": days_to,
                "sort_order": order,
                "is_active": True,
            },
        )
        ranges.append(day_range)
    return ranges


def upsert_rate(season, vehicle_group, day_range, value, note=""):
    rate, _ = VehicleRate.objects.update_or_create(
        season=season,
        vehicle_group=vehicle_group,
        day_range=day_range,
        defaults={
            "daily_rate_gross": Decimal(str(value)),
            "currency": "PLN",
            "is_active": True,
            "note": note,
        },
    )
    return rate


def import_car_free(path):
    supplier = get_supplier("Car Free")
    price_list = upsert_price_list(
        supplier,
        "Car Free customer rates 2026",
        "2026-08-13",
        path.name,
        date(2026, 6, 25),
        None,
    )
    ranges = upsert_ranges(
        price_list,
        [
            ("D1", "1 day", 1, 1),
            ("D2", "2 days", 2, 2),
            ("D3_6", "3-6 days", 3, 6),
            ("D7_14", "7-14 days", 7, 14),
            ("D15_29", "15-29 days", 15, 29),
        ],
    )
    sheet_definitions = [
        ("CarFree Fleet 25.06.2026-15.08", "HIGH_1", "25 Jun - 15 Aug", date(2026, 6, 25), date(2026, 8, 15)),
        ("CarFree Fleet 16.08-31.08", "HIGH_2", "16 Aug - 31 Aug", date(2026, 8, 16), date(2026, 8, 31)),
        ("CarFree Fleet 1.09 - ", "FROM_SEP", "From 1 Sep", date(2026, 9, 1), None),
    ]
    group_map = {
        ("B", "Manual"): "B-MANUAL",
        ("B", "Automatic"): "B-AUTOMATIC",
        ("C Crossover", "Automatic"): "C-CROSSOVER-AUTOMATIC",
        ("C Station Wagon", "Automatic"): "C-STATION-WAGON-AUTOMATIC",
        ("D", "Automatic"): "D-AUTOMATIC",
        ("SUV (TUCSON)", "Automatic"): "SUV-MEDIUM-AUTOMATIC",
        ("SUV", "Automatic"): "SUV-MEDIUM-AUTOMATIC",
        ("SUV 7 Seater", "Automatic"): "SUV-7-SEATER-AUTOMATIC",
        ("E", "Automatic"): "E-AUTOMATIC",
        ("BUS 9 Seater (Toyota Proace)", "Automatic"): "BUS-9-SEATER-AUTOMATIC",
    }
    workbook = load_workbook(path, read_only=True, data_only=True)
    touched = []
    for sheet_name, season_code, season_name, start, end in sheet_definitions:
        season = upsert_season(price_list, season_code, season_name, start, end)
        for row in workbook[sheet_name].iter_rows(min_row=3, values_only=True):
            if not row[0] or not row[1] or not isinstance(row[2], (int, float)):
                continue
            key = (str(row[0]).strip(), str(row[1]).strip())
            if key not in group_map:
                raise CommandError(f"Unmapped Car Free group: {key}")
            vehicle_group = get_group(supplier, group_map[key])
            for day_range, value in zip(ranges, row[2:7]):
                touched.append(
                    upsert_rate(season, vehicle_group, day_range, value, str(row[7] or "")).id
                )
    price_list_rates = VehicleRate.objects.filter(season__price_list=price_list)
    price_list_rates.exclude(id__in=touched).update(is_active=False)
    return price_list, len(touched)


def import_kaizen(path):
    supplier = get_supplier("Kaizen Rent")
    price_list = upsert_price_list(
        supplier,
        "Kaizen With Comfort Package 2026",
        "2026-COMFORT",
        path.name,
        date(2026, 1, 1),
        date(2026, 12, 31),
        "Only With Comfort Package is active for Idan customers.",
    )
    ranges = upsert_ranges(
        price_list,
        [
            ("D1", "1 day", 1, 1),
            ("D2", "2 days", 2, 2),
            ("D3_6", "3-6 days", 3, 6),
            ("D7_14", "7-14 days", 7, 14),
            ("D15_PLUS", "15+ days", 15, None),
        ],
    )
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Idan"]
    season_definitions = [
        ("LOW_BEFORE", "Low season", date(2026, 1, 1), date(2026, 6, 23), 67, 92),
        ("HIGH", "High season", date(2026, 6, 24), date(2026, 8, 21), 5, 30),
        ("LOW_AFTER", "Low season", date(2026, 8, 22), date(2026, 12, 31), 67, 92),
    ]
    touched = []
    for season_code, season_name, start, end, row_from, row_to in season_definitions:
        season = upsert_season(
            price_list,
            season_code,
            season_name,
            start,
            end,
            "With Comfort Package",
        )
        for row in sheet.iter_rows(min_row=row_from, max_row=row_to, values_only=True):
            acriss_value = row[1]
            if not acriss_value or not isinstance(row[4], (int, float)):
                continue
            codes = [code.strip() for code in str(acriss_value).split(";")]
            for code in codes:
                vehicle_group = get_group(supplier, code)
                for day_range, value in zip(ranges, row[4:9]):
                    touched.append(
                        upsert_rate(
                            season,
                            vehicle_group,
                            day_range,
                            value,
                            "With Comfort Package",
                        ).id
                    )
    VehicleRate.objects.filter(season__price_list=price_list).exclude(
        id__in=touched
    ).update(is_active=False)
    return price_list, len(touched)


def import_one_rent(path):
    supplier = get_supplier("One Rent")
    price_list = upsert_price_list(
        supplier,
        "One Rent customer rates 2026",
        "2026",
        path.name,
        date(2026, 1, 1),
        date(2027, 4, 30),
    )
    ranges = upsert_ranges(
        price_list,
        [
            ("D1_2", "1-2 days", 1, 2),
            ("D3_6", "3-6 days", 3, 6),
            ("D7_14", "7-14 days", 7, 14),
            ("D15_PLUS", "15+ days", 15, None),
        ],
    )
    workbook = load_workbook(path, read_only=True, data_only=True)
    season_definitions = [
        ("LOW_BEFORE", "Low season", "Low season", date(2026, 1, 1), date(2026, 4, 30), 5, 24),
        ("HIGH", "High season", "High season", date(2026, 5, 1), date(2026, 9, 30), 4, 23),
        ("LOW_AFTER", "Low season", "Low season", date(2026, 10, 1), date(2027, 4, 30), 5, 24),
    ]
    touched = []
    for season_code, season_name, sheet_name, start, end, row_from, row_to in season_definitions:
        season = upsert_season(price_list, season_code, season_name, start, end)
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=row_from, max_row=row_to, values_only=True):
            code = row[2]
            if not code or not isinstance(row[4], (int, float)):
                continue
            vehicle_group = get_group(supplier, str(code).strip())
            for day_range, value in zip(ranges, row[4:8]):
                touched.append(upsert_rate(season, vehicle_group, day_range, value).id)
    VehicleRate.objects.filter(season__price_list=price_list).exclude(
        id__in=touched
    ).update(is_active=False)
    return price_list, len(touched)


class Command(BaseCommand):
    help = "Import versioned gross vehicle rates from supplier Excel price lists."

    @transaction.atomic
    def handle(self, *args, **options):
        sources = [
            (PRICE_LIST_DIR / "CarFree 13.08.2026.xlsx", import_car_free),
            (PRICE_LIST_DIR / "Kaizen Rent.xlsx", import_kaizen),
            (PRICE_LIST_DIR / "One Rent 2026.xlsx", import_one_rent),
        ]
        total = 0
        for path, importer in sources:
            if not path.exists():
                raise CommandError(f"Price list not found: {path}")
            price_list, count = importer(path)
            total += count
            self.stdout.write(f"{price_list.supplier.supplier_name}: {count} rates")
        self.stdout.write(self.style.SUCCESS(f"Import complete: {total} active rates"))
